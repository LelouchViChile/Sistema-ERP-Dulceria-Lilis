from datetime import datetime
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator
from lilis_erp.roles import require_roles
from django.views.decorators.http import require_POST
import json

from .models import MovimientoInventario, Producto, Proveedor

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def _build_transaction_q(q: str):
    """Construye un Q de búsqueda para movimientos de inventario."""
    q = (q or "").strip()
    if not q:
        return Q()
    return (
        Q(producto__nombre__icontains=q) |
        Q(producto__sku__icontains=q) |
        Q(tipo__icontains=q) |
        Q(proveedor__razon_social__icontains=q) |
        Q(lote__icontains=q) |
        Q(serie__icontains=q) |
        Q(creado_por__username__icontains=q)
    )


@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO", "VENTAS", "COMPRAS")
def gestion_transacciones(request):
    """
    Vista para listar, filtrar, ordenar y exportar movimientos de inventario.
    """
    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', '-fecha')
    ver = request.GET.get('ver', 'todos')
    export = request.GET.get('export', '')

    valid_sort_fields = ['fecha', '-fecha', 'producto__nombre', '-producto__nombre', 'tipo', '-tipo']
    if sort_by not in valid_sort_fields:
        sort_by = '-fecha'

    qs = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'bodega_origen', 'bodega_destino', 'creado_por'
    ).all()

    # Filtro por tipo de movimiento
    if ver == 'ingresos':
        qs = qs.filter(tipo__in=['Ingreso', 'Devolución'])
    elif ver == 'salidas':
        qs = qs.filter(tipo='Salida')
    elif ver == 'ajustes':
        qs = qs.filter(tipo='Ajuste')

    if query:
        qs = qs.filter(_build_transaction_q(query))

    qs = qs.order_by(sort_by)

    if export == "xlsx":
        if Workbook is None:
            return HttpResponse("Falta dependencia: instala openpyxl (pip install openpyxl)", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        headers = [
            "ID", "Fecha", "Tipo", "Producto", "SKU", "Cantidad", "Bodega Origen", "Bodega Destino",
            "Proveedor", "Lote", "Serie", "Vencimiento", "Usuario", "Observación"
        ]
        ws.append(headers)

        for m in qs:
            ws.append([
                m.id,
                m.fecha.strftime('%Y-%m-%d %H:%M'),
                m.get_tipo_display(),
                m.producto.nombre,
                m.producto.sku,
                m.cantidad,
                m.bodega_origen.nombre if m.bodega_origen else "-",
                m.bodega_destino.nombre if m.bodega_destino else "-",
                m.proveedor.razon_social if m.proveedor else "-",
                m.lote or "-",
                m.serie or "-",
                m.fecha_vencimiento.strftime('%Y-%m-%d') if m.fecha_vencimiento else "-",
                m.creado_por.username if m.creado_por else "Sistema",
                m.observacion
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"movimientos_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "movimientos": page_obj.object_list,
        "page_obj": page_obj,
        "query": query,
        "sort_by": sort_by,
        "ver": ver,
    }
    return render(request, "gestion_transacciones.html", context)


@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
@require_POST
def crear_transaccion(request):
    """
    Crea una nueva transacción.
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Payload inválido"}, status=400)

    # Validaciones básicas
    errors = {}
    tipo = (data.get("tipo") or "").strip()
    fecha_str = (data.get("fecha") or "").strip()
    cantidad_str = data.get("cantidad")
    producto_text = (data.get("producto_text") or "").strip()
    proveedor_text = (data.get("proveedor_text") or "").strip()

    if not tipo:
        errors["tipo"] = "Tipo requerido."
    if not fecha_str:
        errors["fecha"] = "Fecha requerida."
    if not producto_text:
        errors["producto_text"] = "Producto requerido."

    try:
        cantidad = float(cantidad_str)
        if cantidad < 0:
            errors["cantidad"] = "La cantidad debe ser mayor a cero."
    except Exception:
        errors["cantidad"] = "Cantidad inválida."

    # Validar proveedor para ingresos
    if tipo == "Ingreso" and not proveedor_text:
        errors["proveedor_text"] = "Proveedor es requerido para ingresos."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    # Buscar producto y proveedor
    producto = (Producto.objects.filter(sku=producto_text).first() or
                Producto.objects.filter(nombre__iexact=producto_text).first())
    if not producto:
        return JsonResponse({"ok": False, "errors": {"producto_text": "Producto no encontrado."}}, status=400)

    proveedor = None
    if proveedor_text:
        proveedor = (Proveedor.objects.filter(rut_nif=proveedor_text).first() or
                     Proveedor.objects.filter(razon_social__iexact=proveedor_text).first())
        if not proveedor and tipo == "Ingreso":
            return JsonResponse({"ok": False, "errors": {"proveedor_text": "Proveedor no encontrado."}}, status=400)

    try:
        with transaction.atomic():
            mov = MovimientoInventario.objects.create(
                fecha=fecha_str,
                tipo=tipo,
                producto=producto,
                proveedor=proveedor,
                cantidad=cantidad,
                lote=data.get("lote", ""),
                serie=data.get("serie", ""),
                fecha_vencimiento=data.get("vencimiento") or None,
                doc_ref=data.get("doc_ref", ""),
                motivo=data.get("motivo", ""),
                observacion=data.get("observaciones", ""),
                creado_por=request.user
            )
            # Aquí iría la lógica para actualizar el stock del producto si es necesario
            # producto.stock += cantidad (o -=)
            # producto.save()
        return JsonResponse({"ok": True, "id": mov.id})
    except Exception as e:
        return JsonResponse({"ok": False, "errors": {"__all__": f"Error inesperado: {e}"}}, status=500)


@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
def editar_transaccion(request, mov_id):
    try:
        mov = MovimientoInventario.objects.get(id=mov_id)
    except MovimientoInventario.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Movimiento no encontrado"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "ok": True,
            "movimiento": {
                "id": mov.id,
                "fecha": mov.fecha.isoformat() if mov.fecha else "",
                "tipo": mov.tipo,
                "producto": getattr(mov.producto, "id", None),
                "proveedor": getattr(mov.proveedor, "id", None),
                "cantidad": mov.cantidad,
            }
        })

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Payload inválido"}, status=400)

    mov.tipo = data.get("tipo", mov.tipo)
    mov.cantidad = data.get("cantidad", mov.cantidad)
    mov.save()

    return JsonResponse({"ok": True})


@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
@require_POST
def eliminar_transaccion(request, mov_id):
    try:
        MovimientoInventario.objects.get(id=mov_id).delete()
        return JsonResponse({"ok": True})
    except MovimientoInventario.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Movimiento no encontrado"}, status=404)
    
@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO", "VENTAS")
def export_xlsx(request):
    """
    Exporta los movimientos de inventario a un archivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = [
        "Fecha", "Tipo", "Producto", "Proveedor", "Cantidad",
        "Usuario", "Lote", "Serie", "Vencimiento", "Doc Ref"
    ]
    ws.append(headers)

    movimientos = MovimientoInventario.objects.select_related("producto", "proveedor").order_by("-fecha")

    for m in movimientos:
        ws.append([
            m.fecha.strftime("%Y-%m-%d"),
            dict(MovimientoInventario.TIPOS).get(m.tipo, m.tipo),
            getattr(m.producto, "nombre", ""),
            getattr(m.proveedor, "razon_social", ""),
            m.cantidad,
            m.usuario.username,
            m.lote or "",
            m.serie or "",
            m.vencimiento.strftime("%Y-%m-%d") if m.vencimiento else "",
            m.doc_ref or ""
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="movimientos.xlsx"'
    wb.save(response)
    return response
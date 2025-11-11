from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction, models
import json
from django.shortcuts import get_object_or_404

from lilis_erp.roles import require_roles

# ===== Excel (opcional) =====
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # se notificará si falta la librería

# ===== Modelo correcto =====
# Tu modelo se llama "Producto", lo importamos y lo aliasamos como Product para mantener el resto del código limpio.
from .models import Producto as Product

from apps.transactional.models import Bodega
from .models import Categoria
# -------- Helpers --------
ALLOWED_SORT_FIELDS = {"id", "sku", "nombre", "categoria", "stock"}


def _display_categoria(obj):
    """
    Devuelve un string para la categoría sin forzar el acceso a la FK
    si está vacía o si la FK quedó apuntando a un ID inexistente.
    """
    cat_id = getattr(obj, "categoria_id", None)
    if not cat_id:
        return ""

    try:
        cat = getattr(obj, "categoria", None)
        if cat is None:
            return ""
        return getattr(cat, "nombre", str(cat)) or ""
    except Exception:
        return ""


def _qs_to_dicts(qs):
    """
    Convierte un queryset de Product a lista de dicts con los campos usados en la tabla/buscador.
    """
    data = []
    for p in qs:
        # Usamos el stock total anotado para eficiencia.
        total_stock = getattr(p, 'stock_total', None)
        if total_stock is None:
            # Fallback por si la anotación no está presente
            total_stock = p.stocks.aggregate(total=models.Sum('cantidad'))['total'] or 0
        data.append({
            "id": p.id,
            "sku": getattr(p, "sku", ""),
            "nombre": getattr(p, "nombre", getattr(p, "name", "")),
            "categoria": _display_categoria(p),
            "stock": total_stock if total_stock is not None else 0,
        })
    return data


def _build_search_q(q: str):
    q = (q or "").strip()
    if not q:
        return Q()

    # Expresión base para buscar en los campos de texto más comunes.
    expr = (
        Q(sku__icontains=q) |
        Q(nombre__icontains=q) |
        Q(marca__icontains=q) | # Búsqueda por marca
        Q(ean_upc__icontains=q) | # Búsqueda por código de barras
        Q(categoria__nombre__icontains=q) # Búsqueda por nombre de categoría
    )

    # Añadir búsqueda por ID si el término es un número.
    try:
        expr |= Q(id=int(q))
    except ValueError:
        pass

    # Stock total exacto si escribe un número
    try:
        expr |= Q(stock_total=int(q))
    except Exception:
        pass

    return expr

# ===================== VISTAS =====================

@login_required
@require_roles("ADMIN", "INVENTARIO", "PRODUCCION", "VENTAS")
def product_list_view(request):
    """
    Listado HTML con:
    - Filtro por q (sku, nombre, categoria*)
    - Orden por sort (id, sku, nombre, categoria, stock)
    - Paginación (10)
    - Exportar Excel (?export=xlsx)
    """
    query = (request.GET.get("q") or "").strip()
    sort_by = (request.GET.get("sort") or "sku").strip()
    export = (request.GET.get("export") or "").strip()

    qs = Product.objects.all().annotate(stock_total=Sum('stocks__cantidad'))

    if query:
        # La búsqueda por stock es especial porque es un campo calculado (anotado).
        # Si el término de búsqueda es numérico, intentamos filtrar por stock.
        try:
            stock_query = int(query)
            # Anotamos el stock total y filtramos por él.
            qs = qs.annotate(total_stock=models.Sum('stocks__cantidad')).filter(
                _build_search_q(query) | Q(total_stock=stock_query)
            )
        except (ValueError, TypeError):
            # Si no es un número, filtramos por los campos de texto habituales.
            qs = qs.filter(_build_search_q(query))

    reverse = sort_by.startswith("-")
    key = sort_by.lstrip("-")
    if key not in ALLOWED_SORT_FIELDS:
        key = "sku"
        reverse = False
    
    # Mapeo de campos pseudo (categoria->categoria__nombre, stock->stock_total)
    sort_field = key
    if key == 'categoria':
        sort_field = 'categoria__nombre'
    elif key == 'stock':
        sort_field = 'stock_total'
    ordering = f'-{sort_field}' if reverse else sort_field
    qs = qs.order_by(ordering)

    # Exportar a Excel
    if export == "xlsx":
        if Workbook is None:
            return HttpResponse(
                "Falta dependencia: instala openpyxl (pip install openpyxl)",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        headers = ["ID", "SKU", "Nombre", "Categoría", "Stock"]
        ws.append(headers)
        
        # Usamos la misma función _qs_to_dicts para asegurar consistencia de datos
        productos_data = _qs_to_dicts(qs)
        for p_data in productos_data:
            ws.append([
                p_data["id"],
                p_data["sku"],
                p_data["nombre"],
                p_data["categoria"],
                p_data["stock"],
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"productos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # 🔹 MEJORA: Pasar las categorías a la plantilla para el formulario de creación
    context = {
        "productos": _qs_to_dicts(page_obj.object_list),
        "page_obj": page_obj,
        "query": query,
        "sort_by": sort_by,
        "categorias": Categoria.objects.all(),
        "uom_choices": Product.UOMS,  # Pasar las opciones de UoM a la plantilla
        "bodegas": Bodega.objects.all(),
    }
    return render(request, "productos.html", context)


@login_required
@require_roles("ADMIN", "INVENTARIO", "PRODUCCION", "VENTAS")
def search_products(request):
    """
    Endpoint AJAX (JSON): retorna máx. 10 productos filtrados por 'q'.
    """
    q = (request.GET.get("q") or "").strip()
    qs = Product.objects.all()
    if q:
        qs = qs.filter(_build_search_q(q))

    qs = qs.order_by("id")[:10]
    return JsonResponse({"results": _qs_to_dicts(qs)})

@login_required
@require_roles("ADMIN", "INVENTARIO", "PRODUCCION","VENTAS")
@transaction.atomic
def crear_producto(request):
    """
    Inserta un producto usando la misma lógica de 'transacciones:crear':
    - Recibe JSON (Content-Type: application/json) o form-url-encoded.
    - Valida mínimos.
    - Crea el Producto.
    - Responde JSON con { ok: True } en éxito.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método no permitido."}, status=405)

    # --- 1) Parseo de datos al estilo transacciones (preferir JSON) ---
    is_json = request.headers.get("Content-Type", "").startswith("application/json")
    if is_json:
        try:
            data = json.loads(request.body.decode("utf-8"))
        except Exception:
            return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)
    else:
        data = request.POST

    # --- 2) Extraer campos del FRONT y mapear a tu MODELO real ---
    # Campos del form (nombres del template):
    sku           = (data.get("sku") or "").strip().upper()
    nombre        = (data.get("nombre") or "").strip()
    categoria_id  = (data.get("categoria") or "").strip() or None
    descripcion   = (data.get("descripcion") or "").strip()
    precio_compra = (data.get("precio_compra") or "").strip()
    precio_venta  = (data.get("precio_venta") or "").strip()
    marca         = (data.get("marca") or "").strip()
    ean_upc       = (data.get("codigo_barras") or "").strip()
    ubicacion_id  = (data.get("ubicacion") or "").strip()
    uom           = (data.get("uom") or "UN").strip() # Leer la unidad de medida
    # NOTA: el template tiene campos como 'unidad_medida', 'ubicacion', 'proveedor',
    # 'ultima_entrada', 'imagen', 'notas' que NO existen en el modelo -> los ignoramos
    # o los mapearías a otros campos si los agregas al modelo en el futuro.
    #
    # Mapeo a modelo Producto:
    # - costo_estandar  <- precio_compra
    # - precio_venta    <- precio_venta
    # - stock_minimo    <- stock_min
    # - stock_maximo    <- stock_max
    # - activo          <- estado ("Activo" => True, otro => False)
    stock_min     = (data.get("stock_min") or "").strip()
    stock_max     = (data.get("stock_max") or "").strip()
    estado        = (data.get("estado") or "Activo").strip()


    # --- 3) Validaciones clave (mismo estilo que transacciones) ---
    if not sku or not nombre:
        return JsonResponse({"ok": False, "error": "SKU y Nombre son obligatorios."}, status=400)
    if not categoria_id:
        return JsonResponse({"ok": False, "error": "Selecciona una Categoría."}, status=400)
    if not ubicacion_id:
        return JsonResponse({"ok": False, "error": "Selecciona una Ubicación."}, status=400)

    # numéricos
    def to_decimal(val, default=None):
        if val in (None, ""):
            return default
        try:
            return float(val)
        except Exception:
            return None

    costo_estandar = to_decimal(precio_compra, default=None)
    precio_venta_f = to_decimal(precio_venta, default=None)
    stock_minimo   = to_decimal(stock_min,   default=0)
    stock_maximo   = to_decimal(stock_max,   default=None)

    if precio_venta and precio_venta_f is None:
        return JsonResponse({"ok": False, "error": "Precio de venta inválido."}, status=400)
    if precio_compra and costo_estandar is None:
        return JsonResponse({"ok": False, "error": "Precio de compra inválido."}, status=400)
    if stock_min and stock_minimo is None:
        return JsonResponse({"ok": False, "error": "Stock mínimo inválido."}, status=400)
    if stock_max and stock_maximo is None:
        return JsonResponse({"ok": False, "error": "Stock máximo inválido."}, status=400)

    # categoría existe?
    get_object_or_404(Categoria, id=categoria_id)
    
    # bodega existe?
    get_object_or_404(Bodega, id=ubicacion_id)

    # SKU único
    if Product.objects.filter(sku=sku).exists():
        return JsonResponse({"ok": False, "error": "El SKU ya existe."}, status=400)

    # Activo desde el estado del form
    activo = (estado.upper() == "ACTIVO")

    # --- 4) Crear Producto ---
    try:
        prod = Product.objects.create(
            sku=sku,
            ean_upc=ean_upc or None,
            nombre=nombre,
            descripcion=descripcion,
            categoria_id=categoria_id,
            marca=marca or "",
            # Usar la UoM del formulario para compra y venta
            uom_compra=uom,
            uom_venta=uom,
            factor_conversion=1,
            costo_estandar=costo_estandar if costo_estandar is not None else None,
            precio_venta=precio_venta_f if precio_venta_f is not None else None,
            impuesto_iva=19,  # puedes exponerlo en el form si quieres
            stock_minimo=stock_minimo if stock_minimo is not None else 0,
            stock_maximo=stock_maximo, # Se mantiene la validación de stock_maximo >= stock_minimo en el modelo
            punto_reorden=None,
            perecible=False,
            control_por_lote=False,
            control_por_serie=False,
            url_imagen="",            # (el input del template es file; no lo manejamos por JSON)
            url_ficha_tecnica="",
            activo=activo,
        )
        return JsonResponse({"ok": True, "id": prod.id, "message": "Producto creado correctamente."})
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"No se pudo crear: {e}"}, status=500)


@login_required
@csrf_exempt
def eliminar_producto(request, prod_id):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Método no permitido."}, status=405)

    producto = get_object_or_404(Product, id=prod_id)
    producto.delete()
    return JsonResponse({"status": "ok", "message": "Producto eliminado correctamente."})


@login_required
@require_roles("ADMIN", "INVENTARIO", "PRODUCCION","VENTAS")
@csrf_exempt
def editar_producto(request, prod_id):
    producto = get_object_or_404(Product, id=prod_id)

    if request.method == "GET":
        # Obtenemos todas las categorías para pasarlas al modal
        categorias = list(Categoria.objects.all().values('id', 'nombre'))
        # Solo devolvemos los campos que realmente existen en el modelo
        return JsonResponse({
            "id": producto.id,
            "sku": producto.sku,
            "nombre": producto.nombre,
            "categoria": producto.categoria_id or "",
            "categorias": categorias, # Enviamos la lista de categorías
            "descripcion": producto.descripcion or "",
            "precio_venta": producto.precio_venta or 0,
            "marca": producto.marca or "",
        })

    if request.method == "POST":
        sku = (request.POST.get("sku") or "").strip()
        nombre = (request.POST.get("nombre") or "").strip()
        categoria_id = request.POST.get("categoria") or None
        descripcion = (request.POST.get("descripcion") or "").strip()
        precio_venta = request.POST.get("precio_venta") or 0
        marca = (request.POST.get("marca") or "").strip()

        if not all([sku, nombre]):
            return JsonResponse({"status": "error", "message": "SKU y nombre son obligatorios."})

        if Product.objects.filter(sku=sku).exclude(id=producto.id).exists():
            return JsonResponse({"status": "error", "message": "El SKU ya existe."})

        # 🔹 CORRECCIÓN: Actualizar el objeto 'producto' existente en lugar de crear uno nuevo.
        producto.sku = sku or producto.sku
        producto.nombre = nombre or producto.nombre
        if categoria_id: # Solo actualiza si se recibe un ID de categoría
            producto.categoria_id = categoria_id
        producto.descripcion = descripcion or producto.descripcion
        producto.precio_venta = precio_venta or producto.precio_venta
        producto.marca = marca or producto.marca
        
        # 🔹 CORRECCIÓN: Especificar los campos a actualizar para evitar problemas con campos automáticos.
        update_fields = ['sku', 'nombre', 'descripcion', 'precio_venta', 'marca']
        if categoria_id:
            update_fields.append('categoria_id')
        producto.save(update_fields=update_fields)

        return JsonResponse({"status": "ok", "message": "Producto actualizado correctamente."})

    return JsonResponse({"status": "error", "message": "Método no permitido."}, status=405) 

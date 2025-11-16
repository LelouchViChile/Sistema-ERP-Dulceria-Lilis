from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, HttpRequest
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import transaction
from datetime import datetime

from .models import Usuario
from .utils_invite import invite_user_and_email

# ====== export a Excel (openpyxl) ======
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

def _usuarios_to_excel(queryset):
    if Workbook is None:
        return HttpResponse(
            "Falta dependencia: instala openpyxl (pip install openpyxl)",
            status=500,
            content_type="text/plain; charset=utf-8",
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    headers = ["ID", "Username", "Email", "Nombre", "Apellido", "Teléfono",
               "Rol", "Estado", "Activo", "MFA", "Último acceso", "Creado"]
    ws.append(headers)
    for u in queryset:
        ws.append([
            u.id, u.username, u.email, u.first_name or "", u.last_name or "",
            u.telefono or "", getattr(u, "rol", ""), getattr(u, "estado", ""),
            "Sí" if u.activo else "No",
            "Sí" if getattr(u, "mfa_habilitado", False) else "No",
            u.last_login.strftime("%d/%m/%Y %H:%M") if u.last_login else "",
            u.date_joined.strftime("%d/%m/%Y %H:%M") if u.date_joined else "",
        ])
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def _rol_from_text(q: str):
    q = (q or "").strip().lower()
    mapping = {
        'administrador': 'ADMIN',
        'admin': 'ADMIN',
        'compras': 'COMPRAS',
        'inventario': 'INVENTARIO',
        'produccion': 'PRODUCCION',
        'producción': 'PRODUCCION',
        'ventas': 'VENTAS',
        'finanzas': 'FINANZAS',
        'soporte': 'SOPORTE',
    }
    return mapping.get(q)


# ====== export a Excel (openpyxl) ======


def _es_admin(user):
    return user.is_superuser or getattr(user, "rol", "") == "ADMIN"


@login_required
def gestion_usuarios(request):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', 'id')
    export = request.GET.get('export')
    ver = request.GET.get('ver', 'todos')  # activos | inactivos | todos

    valid_sort_fields = ['id', '-id', 'username', '-username', 'first_name', '-first_name', 'rol', '-rol']
    if sort_by not in valid_sort_fields:
        sort_by = 'id'

    usuarios_list = Usuario.objects.all()

    # Filtro de estado/activo
    if ver == 'inactivos':
        usuarios_list = usuarios_list.filter(Q(estado__in=['inactivo', 'bloqueado']) | Q(activo=False))
    elif ver == 'activos':
        usuarios_list = usuarios_list.filter(estado='activo', activo=True)
    # ver == 'todos' no filtra

    q = (request.GET.get('q') or '').strip()
    expr = Q()
    if q:
        # username, nombre, apellido, email, teléfono
        expr |= Q(username__icontains=q)
        expr |= Q(first_name__icontains=q)
        expr |= Q(last_name__icontains=q)
        expr |= Q(email__icontains=q)
        expr |= Q(telefono__icontains=q)

        # ID exacto si es número
        try:
            expr |= Q(id=int(q))
        except ValueError:
            pass

        # Rol por etiqueta de texto
        rol_code = _rol_from_text(q)
        if rol_code:
            expr |= Q(rol=rol_code)

        usuarios_list = usuarios_list.filter(expr)

    usuarios_list = usuarios_list.order_by(sort_by)

    if export == 'xlsx':
        return _usuarios_to_excel(usuarios_list)
    
    page_number = request.GET.get('page', 1)
    paginator = Paginator(usuarios_list, 10)
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        'usuarios/gestion_usuarios.html',
        {'page_obj': page_obj, 'query': query, 'sort_by': sort_by, 'ver': ver}
    )


@login_required
@transaction.atomic
def crear_usuario(request):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    username = (request.POST.get('username') or '').strip()
    email = (request.POST.get('email') or '').strip()
    nombre = (request.POST.get('first_name') or '').strip()
    apellido = (request.POST.get('last_name') or '').strip()
    telefono = (request.POST.get('telefono') or '').strip()
    rol = (request.POST.get('rol') or '').strip()
    estado = (request.POST.get('estado') or '').strip()
    mfa_habilitado = request.POST.get('mfa_habilitado') == 'on'

    if not all([username, email, nombre, apellido, telefono, rol, estado]):
        return JsonResponse({'status': 'error', 'message': 'Todos los campos son obligatorios.'})

    if Usuario.objects.filter(username=username).exists():
        return JsonResponse({'status': 'error', 'message': 'El nombre de usuario ya existe.'})
    if Usuario.objects.filter(email=email).exists():
        return JsonResponse({'status': 'error', 'message': 'El email ya está registrado.'})
    if Usuario.objects.filter(telefono=telefono).exists():
        return JsonResponse({'status': 'error', 'message': 'El teléfono ya está registrado.'})

    try:
        usuario = Usuario.objects.create(
            username=username,
            email=email,
            first_name=nombre,
            last_name=apellido,
            telefono=telefono,
            rol=rol,
            estado=estado,
            activo=(estado == 'activo'),
            mfa_habilitado=mfa_habilitado,
        )
        invite_user_and_email(usuario)
        return JsonResponse({'status': 'ok', 'message': 'Usuario creado e invitación enviada.'})
    except ValidationError as e:
        return JsonResponse({'status': 'error', 'message': ' '.join(e.messages)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'No se pudo crear el usuario: {e}'}, status=500)

@transaction.atomic
@login_required
@csrf_exempt
def eliminar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    usuario = get_object_or_404(Usuario, id=user_id)

    # No se puede eliminar a sí mismo
    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes eliminar tu propia cuenta.'}, status=403)

    # Solo un superusuario puede eliminar a otro administrador.
    # Un administrador normal (rol=ADMIN) no puede eliminar a otro admin o superuser.
    if _es_admin(usuario) and not request.user.is_superuser:
        return JsonResponse({
            'status': 'error', 'message': 'Solo un Superusuario puede eliminar a otro administrador.'
        }, status=403)

    usuario.delete()
    return JsonResponse({'status': 'ok', 'message': 'Usuario eliminado correctamente.'})


@login_required
@csrf_exempt
def editar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        email    = (request.POST.get('email') or '').strip()
        nombre   = (request.POST.get('first_name') or '').strip()
        apellido = (request.POST.get('last_name') or '').strip()
        telefono = (request.POST.get('telefono') or '').strip()
        rol      = (request.POST.get('rol') or '').strip()
        estado   = (request.POST.get('estado') or '').strip()

        if Usuario.objects.filter(username=username).exclude(id=usuario.id).exists():
            return JsonResponse({'status': 'error', 'message': 'El nombre de usuario ya existe.'})
        if Usuario.objects.filter(email=email).exclude(id=usuario.id).exists():
            return JsonResponse({'status': 'error', 'message': 'El email ya está registrado.'})
        if Usuario.objects.filter(telefono=telefono).exclude(id=usuario.id).exists():
            return JsonResponse({'status': 'error', 'message': 'El teléfono ya está registrado.'})

        usuario.username   = username
        usuario.email      = email
        usuario.first_name = nombre
        usuario.last_name  = apellido
        usuario.telefono   = telefono
        usuario.rol        = rol
        usuario.estado     = estado
        usuario.activo     = (estado == 'activo')
        usuario.save()
        return JsonResponse({'status': 'ok', 'message': 'Usuario actualizado correctamente.'})

    return JsonResponse({
        'id': usuario.id,
        'username': usuario.username,
        'email': usuario.email,
        'first_name': usuario.first_name,
        'last_name': usuario.last_name,
        'telefono': usuario.telefono,
        'rol': usuario.rol,
        'estado': usuario.estado,
        'activo': usuario.activo,
    })


# ---- Nuevos endpoints: desactivar / reactivar (solo ADMIN) ----
@login_required
@csrf_exempt
def desactivar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    # No se puede desactivar a sí mismo
    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes desactivar tu propia cuenta.'}, status=403)

    # Un administrador no puede desactivar a un superusuario.
    # Solo un superusuario puede gestionar a otro superusuario.
    if usuario.is_superuser and not request.user.is_superuser:
        return JsonResponse({
            'status': 'error', 'message': 'No tienes permisos para desactivar a un Superusuario.'
        }, status=403)

    usuario.estado = 'inactivo'
    usuario.activo = False
    usuario.save(update_fields=['estado', 'activo'])
    return JsonResponse({'status': 'ok', 'message': 'Usuario desactivado.'})


@login_required
@csrf_exempt
def reactivar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    # No se puede reactivar a sí mismo (caso improbable, pero seguro)
    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'Acción no permitida sobre tu propia cuenta.'}, status=403)

    # Un administrador no puede reactivar a un superusuario.
    # Solo un superusuario puede gestionar a otro superusuario.
    if usuario.is_superuser and not request.user.is_superuser:
        return JsonResponse({
            'status': 'error', 'message': 'No tienes permisos para reactivar a un Superusuario.'
        }, status=403)

    usuario.estado = 'activo'
    usuario.activo = True
    usuario.save(update_fields=['estado', 'activo'])
    return JsonResponse({'status': 'ok', 'message': 'Usuario reactivado.'})


@login_required
@csrf_exempt
def reiniciar_clave(request: HttpRequest, user_id: int):
    """
    Reinicia la clave de un usuario y le envía un correo de invitación
    para que establezca una nueva, reutilizando la lógica de creación.
    """
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    usuario = get_object_or_404(Usuario, id=user_id)
    
    invite_user_and_email(usuario, source='reset')
    
    # Invalidar todas las sesiones existentes para este usuario.
    # La sesión actual del admin no se ve afectada.
    from django.contrib.auth import update_session_auth_hash
    update_session_auth_hash(request, usuario)

    # Si el admin se reinicia su propia clave, lo deslogueamos.
    force_logout = False
    if request.user.id == usuario.id:
        from django.contrib.auth import logout
        logout(request)
        force_logout = True
    
    return JsonResponse({
        'status': 'ok',
        'message': f'Se ha enviado un correo para el reinicio de clave a {usuario.email}.',
        'force_logout': force_logout,
    })

@login_required
@csrf_exempt
def bloquear_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    # No se puede bloquear a sí mismo
    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes bloquear tu propia cuenta.'}, status=403)

    usuario.estado = 'bloqueado'
    usuario.activo = False
    usuario.save(update_fields=['estado', 'activo'])
    return JsonResponse({'status': 'ok', 'message': 'Usuario bloqueado.'})

@login_required
@csrf_exempt
def desbloquear_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes desbloquear tu propia cuenta.'}, status=403)

    usuario.estado = 'activo'
    usuario.activo = True
    usuario.save(update_fields=['estado', 'activo'])
    return JsonResponse({'status': 'ok', 'message': 'Usuario desbloqueado.'})
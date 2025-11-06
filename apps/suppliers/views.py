from datetime import datetime
import json
import re

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST

from lilis_erp.roles import require_roles

# Ajusta si tus nombres de modelos difieren
from apps.suppliers.models import Proveedor, ProveedorProducto
from apps.products.models import Producto

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# -------------------------- Helpers --------------------------

def _supplier_to_dict(s: Proveedor):
    estado = getattr(s, "estado", "")
    if not estado:
        estado = "Activo" if getattr(s, "activo", False) else "Inactivo"
    return {
        "id": s.id,
        "rut": getattr(s, "rut_nif", ""),
        "razon_social": getattr(s, "razon_social", ""),
        "estado": estado,
        "email": getattr(s, "email", ""),
    }


def _rel_to_dict(rel: ProveedorProducto):
    pvd = rel.proveedor
    pro = rel.producto
    return {
        "id": rel.id,
        "proveedor": getattr(pvd, "razon_social", ""),
        "rut": getattr(pvd, "rut_nif", ""),
        "producto": getattr(pro, "nombre", ""),
        "sku": getattr(pro, "sku", ""),
        "preferente": bool(getattr(rel, "preferente", False)),
        "lead_time": getattr(rel, "lead_time_dias", 0) or 0,
        "costo": getattr(rel, "costo", 0) or 0,
        "minimo_lote": getattr(rel, "minimo_lote", 0) or 0,
        "descuento_porcentaje": getattr(rel, "descuento_porcentaje", 0) or 0,
    }


def _build_supplier_q(q: str):
    q = (q or "").strip()
    if not q:
        return Q()
    return (Q(rut_nif__icontains=q) |
            Q(razon_social__icontains=q) |
            Q(email__icontains=q))


def _build_relation_q(q: str):
    q = (q or "").strip()
    if not q:
        return Q()
    return (Q(proveedor__rut_nif__icontains=q) |
            Q(proveedor__razon_social__icontains=q) |
            Q(producto__sku__icontains=q) |
            Q(producto__nombre__icontains=q))


def _valid_email(s: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", s or ""))


# ---------------------------- Vistas ----------------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def supplier_list_view(request):
    q = (request.GET.get("q") or "").strip()

    qs = Proveedor.objects.all().order_by("id")
    if q:
        qs = qs.filter(_build_supplier_q(q))

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "proveedores": [_supplier_to_dict(s) for s in page_obj.object_list],
        "page_obj": page_obj,
        "query": q,
    }
    return render(request, "gestion_proveedores.html", context)


# ------------------------- AJAX: CREATE -------------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def create_supplier(request):
    """
    Crea/actualiza proveedor por RUT. Si existe, actualiza datos básicos.
    Validación:
      - rut_nif requerido y único
      - razon_social requerida
      - email requerido y válido
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "errors": {"__all__": "Payload inválido"}}, status=400)

    errors = {}
    rut = (data.get("rut_nif") or "").strip()
    razon = (data.get("razon_social") or "").strip()
    email = (data.get("email") or "").strip()
    telefono = (data.get("telefono") or "").strip()
    web = (data.get("sitio_web") or "").strip()
    nombre_fantasia = (data.get("nombre_fantasia") or "").strip()

    # comerciales
    try:
        plazo = int(data.get("plazos_pago_dias") or 0)
    except Exception:
        plazo = -1
    moneda = (data.get("moneda") or "CLP").strip()
    try:
        descuento = float(data.get("descuento_porcentaje") or 0)
    except Exception:
        descuento = -1

    # Validaciones
    if not re.match(r"^[0-9kK.\-]{7,20}$", rut or ""):
        errors["rut_nif"] = "RUT/NIF obligatorio y válido."

    if not razon:
        errors["razon_social"] = "Razón social obligatoria."

    if not _valid_email(email):
        errors["email"] = "Email obligatorio y válido."

    if telefono and not re.match(r"^[0-9+()\-\s]{6,30}$", telefono):
        errors["telefono"] = "Formato de teléfono inválido."

    if web and not re.match(r"^https?://", web, flags=re.I):
        errors["sitio_web"] = "Debe comenzar con http:// o https://"

    if not (0 <= plazo <= 365):
        errors["plazos_pago_dias"] = "Plazo de 0 a 365."

    if not (0 <= descuento <= 100):
        errors["descuento_porcentaje"] = "Descuento 0 a 100%."

    # Duplicado
    if not errors:
        existente = Proveedor.objects.filter(rut_nif=rut).first()
        if existente and str(existente.id) != str(data.get("id") or ""):
            errors["rut_nif"] = "Ya existe un proveedor con este RUT/NIF."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    # Crear o actualizar
    with transaction.atomic():
        proveedor, _created = Proveedor.objects.update_or_create(
            rut_nif=rut,
            defaults={
                "razon_social": razon,
                "email": email,
                "telefono": telefono,
                "sitio_web": web,
                "nombre_fantasia": nombre_fantasia,
                "plazos_pago_dias": plazo,
                "moneda": moneda or "CLP",
                "descuento_porcentaje": descuento,
                "activo": True,
            }
        )

    return JsonResponse({"ok": True, "id": proveedor.id})


@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def create_relation(request):
    """
    Crea la relación proveedor–producto:
      - requiere rut_nif de proveedor existente
      - busca producto por SKU exacto o por nombre (primer match)
      - valida no negativos y rangos
      - si ya existe la relación para el par (proveedor, producto), actualiza
    Campos aceptados:
      rut_nif, sku_or_name, preferente(bool), lead_time_dias(int 0-365),
      costo(float>=0), minimo_lote(int>=0 opc), descuento_porcentaje(0-100)
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "errors": {"__all__": "Payload inválido"}}, status=400)

    errors = {}

    rut = (data.get("rut_nif") or "").strip()
    sku_or_name = (data.get("sku_or_name") or "").strip()
    preferente = bool(data.get("preferente"))
    try:
        lead = int(data.get("lead_time_dias") or 0)
    except Exception:
        lead = -1
    try:
        costo = float(data.get("costo") or 0)
    except Exception:
        costo = -1
    try:
        minimo_lote = int(data.get("minimo_lote") or 0)
    except Exception:
        minimo_lote = -1
    try:
        desc = float(data.get("descuento_porcentaje") or 0)
    except Exception:
        desc = -1

    if not rut:
        errors["rut_nif"] = "Debe indicar el RUT/NIF del proveedor (pestaña 1)."
    if not sku_or_name:
        errors["sku_or_name"] = "Ingrese SKU o nombre de producto."
    if not (0 <= lead <= 365):
        errors["lead_time_dias"] = "Lead time 0 a 365."
    if not (costo >= 0):
        errors["costo"] = "Costo no puede ser negativo."
    if not (minimo_lote >= 0):
        errors["minimo_lote"] = "Mínimo lote no puede ser negativo."
    if not (0 <= desc <= 100):
        errors["descuento_porcentaje"] = "Descuento 0 a 100%."

    proveedor = None
    producto = None

    if not errors:
        proveedor = Proveedor.objects.filter(rut_nif=rut).first()
        if not proveedor:
            errors["rut_nif"] = "Proveedor no existe. Guárdalo primero."

    if not errors:
        # buscar producto por SKU exacto o por nombre (primer resultado)
        producto = (Producto.objects.filter(sku=sku_or_name).first() or
                    Producto.objects.filter(nombre__icontains=sku_or_name).order_by("id").first())
        if not producto:
            errors["sku_or_name"] = "Producto no encontrado."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    with transaction.atomic():
        rel, _created = ProveedorProducto.objects.update_or_create(
            proveedor=proveedor,
            producto=producto,
            defaults={
                "preferente": preferente,
                "lead_time_dias": lead,
                "costo": costo,
                "minimo_lote": minimo_lote,
                "descuento_porcentaje": desc,
            }
        )

    return JsonResponse({"ok": True, "id": rel.id})


# --------------------- AJAX: SEARCH / EXPORT ---------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def search_suppliers(request):
    q = (request.GET.get("q") or "").strip()
    qs = Proveedor.objects.all()
    if q:
        qs = qs.filter(_build_supplier_q(q))
    qs = qs.order_by("id")[:10]
    return JsonResponse({"results": [_supplier_to_dict(s) for s in qs]})


@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def relations_search(request):
    q = (request.GET.get("q") or "").strip()
    qs = (ProveedorProducto.objects
          .select_related("proveedor", "producto"))
    if q:
        qs = qs.filter(_build_relation_q(q))
    qs = qs.order_by("id")[:10]
    return JsonResponse({"results": [_rel_to_dict(r) for r in qs]})


@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def relations_export(request):
    if Workbook is None:
        return HttpResponse(
            "Falta dependencia: instala openpyxl (pip install openpyxl)",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    q = (request.GET.get("q") or "").strip()
    qs = (ProveedorProducto.objects
          .select_related("proveedor", "producto"))
    if q:
        qs = qs.filter(_build_relation_q(q))
    qs = qs.order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Relaciones"
    headers = ["ID", "Proveedor", "RUT/NIF", "Producto", "SKU",
               "Preferente", "Lead time (d)", "Costo", "Mínimo lote", "Descuento (%)"]
    ws.append(headers)

    for r in qs:
        d = _rel_to_dict(r)
        ws.append([
            d["id"], d["proveedor"], d["rut"], d["producto"], d["sku"],
            "Sí" if d["preferente"] else "No",
            d["lead_time"], d["costo"], d["minimo_lote"], d["descuento_porcentaje"]
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"relaciones_proveedor_producto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ---------------------- EDITAR Y ELIMINAR ----------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def editar_proveedor(request, supplier_id):
    try:
        proveedor = Proveedor.objects.get(id=supplier_id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": proveedor.id,
            "rut_nif": proveedor.rut_nif,
            "razon_social": proveedor.razon_social,
            "nombre_fantasia": proveedor.nombre_fantasia or "",
            "email": proveedor.email,
            "telefono": proveedor.telefono or "",
            "sitio_web": proveedor.sitio_web or "",
            "plazos_pago_dias": proveedor.plazos_pago_dias or 0,
            "moneda": proveedor.moneda or "CLP",
            "descuento_porcentaje": proveedor.descuento_porcentaje or 0,
        })

    elif request.method == "POST":
        data = request.POST
        proveedor.rut_nif = data.get("rut_nif", proveedor.rut_nif)
        proveedor.razon_social = data.get("razon_social", proveedor.razon_social)
        proveedor.nombre_fantasia = data.get("nombre_fantasia", proveedor.nombre_fantasia)
        proveedor.email = data.get("email", proveedor.email)
        proveedor.telefono = data.get("telefono", proveedor.telefono)
        proveedor.sitio_web = data.get("sitio_web", proveedor.sitio_web)
        proveedor.plazos_pago_dias = data.get("plazos_pago_dias", proveedor.plazos_pago_dias)
        proveedor.moneda = data.get("moneda", proveedor.moneda)
        proveedor.descuento_porcentaje = data.get("descuento_porcentaje", proveedor.descuento_porcentaje)
        proveedor.save()
        return JsonResponse({"status": "ok", "message": "Proveedor actualizado correctamente"})
    else:
        return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)


@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def eliminar_proveedor(request, supplier_id):
    try:
        proveedor = Proveedor.objects.get(id=supplier_id)
        proveedor.delete()
        return JsonResponse({"status": "ok", "message": "Proveedor eliminado correctamente"})
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)